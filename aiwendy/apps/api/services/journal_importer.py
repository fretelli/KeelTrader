"""Import trading journals from CSV/XLSX with user-provided column mapping."""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

MAX_IMPORT_ROWS = 5000
MAX_PREVIEW_ROWS = 50


@dataclass(frozen=True)
class ParsedTable:
    columns: List[str]
    rows: List[Dict[str, Any]]
    warnings: List[str]


def _decode_bytes(content: bytes) -> Tuple[str, List[str]]:
    warnings: List[str] = []
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return content.decode(encoding), warnings
        except UnicodeDecodeError:
            continue
    warnings.append("Failed to decode as UTF-8/GB18030; used replacement decoding.")
    return content.decode("utf-8", errors="replace"), warnings


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def _normalize_header(value: str) -> str:
    return "".join(ch for ch in value.strip().lower() if ch.isalnum())


def _parse_csv(content: bytes, max_rows: int) -> ParsedTable:
    text, warnings = _decode_bytes(content)
    sample = text[:8192]

    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t", "|"])
        delimiter = dialect.delimiter
    except Exception:
        warnings.append("CSV delimiter detection failed; defaulted to ','.")

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows_iter = iter(reader)

    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        return ParsedTable(columns=[], rows=[], warnings=["Empty file."])

    headers: List[str] = []
    seen: Dict[str, int] = {}
    for idx, header in enumerate(raw_headers):
        base = _stringify(header) or f"column_{idx + 1}"
        name = base
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[base]}"
        else:
            seen[name] = 1
        headers.append(name)

    parsed_rows: List[Dict[str, Any]] = []
    for i, row in enumerate(rows_iter):
        if i >= max_rows:
            warnings.append(
                f"Reached max rows limit ({max_rows}); remaining rows ignored."
            )
            break
        values = [_stringify(v) for v in row]
        values += [""] * max(0, len(headers) - len(values))
        parsed_rows.append({headers[j]: values[j] for j in range(len(headers))})

    return ParsedTable(columns=headers, rows=parsed_rows, warnings=warnings)


def _parse_xlsx(content: bytes, max_rows: int) -> ParsedTable:
    warnings: List[str] = []
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration:
        return ParsedTable(columns=[], rows=[], warnings=["Empty file."])

    headers: List[str] = []
    seen: Dict[str, int] = {}
    for idx, header in enumerate(raw_headers):
        base = _stringify(header) or f"column_{idx + 1}"
        name = base
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[base]}"
        else:
            seen[name] = 1
        headers.append(name)

    parsed_rows: List[Dict[str, Any]] = []
    for i, row in enumerate(rows):
        if i >= max_rows:
            warnings.append(
                f"Reached max rows limit ({max_rows}); remaining rows ignored."
            )
            break
        values = [_stringify(v) for v in row]
        values += [""] * max(0, len(headers) - len(values))
        parsed_rows.append({headers[j]: values[j] for j in range(len(headers))})

    return ParsedTable(columns=headers, rows=parsed_rows, warnings=warnings)


def parse_tabular_file(filename: str, content: bytes, max_rows: int) -> ParsedTable:
    if max_rows <= 0:
        return ParsedTable(columns=[], rows=[], warnings=["Invalid max_rows."])

    ext = (filename or "").lower().rsplit(".", 1)[-1] if "." in (filename or "") else ""
    if ext in {"xlsx"}:
        return _parse_xlsx(content, max_rows)
    if ext in {"csv"}:
        return _parse_csv(content, max_rows)
    # Best effort: try CSV first.
    parsed = _parse_csv(content, max_rows)
    parsed.warnings.append("Unknown extension; attempted CSV parsing.")
    return parsed


def suggest_mapping(columns: Iterable[str]) -> Dict[str, Optional[str]]:
    cols = list(columns)
    normalized = {_normalize_header(c): c for c in cols}

    def pick(*candidates: str) -> Optional[str]:
        for cand in candidates:
            key = _normalize_header(cand)
            if key in normalized:
                return normalized[key]
        for col in cols:
            col_key = _normalize_header(col)
            for cand in candidates:
                cand_key = _normalize_header(cand)
                if cand_key and cand_key in col_key:
                    return col
        return None

    return {
        "symbol": pick(
            "symbol", "ticker", "code", "instrument", "标的", "合约", "币种", "品种"
        ),
        "direction": pick(
            "direction", "side", "buy/sell", "buysell", "买卖", "方向", "开仓方向"
        ),
        "trade_date": pick(
            "trade_date",
            "date",
            "datetime",
            "time",
            "开仓时间",
            "成交时间",
            "日期",
            "时间",
        ),
        "entry_price": pick(
            "entry_price",
            "entry",
            "open_price",
            "openprice",
            "开仓价",
            "入场价",
            "买入价",
        ),
        "exit_price": pick(
            "exit_price",
            "exit",
            "close_price",
            "closeprice",
            "平仓价",
            "出场价",
            "卖出价",
        ),
        "position_size": pick(
            "position_size", "quantity", "qty", "size", "amount", "数量", "手数"
        ),
        "pnl_amount": pick(
            "pnl", "profit", "profit_loss", "pnl_amount", "盈亏", "利润"
        ),
        "notes": pick("notes", "note", "comment", "备注", "说明"),
    }


def parse_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _stringify(value)
    if not text:
        return None
    text = (
        text.replace(",", "")
        .replace("$", "")
        .replace("¥", "")
        .replace("￥", "")
        .replace("%", "")
    ).strip()
    try:
        return float(text)
    except ValueError:
        return None


def parse_datetime(value: Any) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = _stringify(value)
    if not text:
        return None

    # Fast path: ISO 8601
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        pass

    patterns = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
    ]
    for pattern in patterns:
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    return None


def parse_direction(value: Any) -> Optional[str]:
    text = _stringify(value).lower()
    if not text:
        return None
    normalized = text.replace(" ", "").replace("_", "").replace("-", "")

    long_values = {"long", "buy", "b", "多", "做多", "买", "买入", "开多"}
    short_values = {"short", "sell", "s", "空", "做空", "卖", "卖出", "开空"}

    if normalized in long_values:
        return "long"
    if normalized in short_values:
        return "short"

    if normalized.startswith("buy"):
        return "long"
    if normalized.startswith("sell"):
        return "short"

    return None


def build_journal_payload(
    row: Dict[str, Any],
    mapping: Dict[str, str],
    *,
    project_id: Optional[str] = None,
) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    def get(col: Optional[str]) -> Any:
        if not col:
            return None
        return row.get(col)

    symbol = _stringify(get(mapping.get("symbol")))
    if not symbol:
        return None, "Missing symbol"

    direction = parse_direction(get(mapping.get("direction")))
    if not direction:
        return None, "Missing/invalid direction"

    trade_date = parse_datetime(get(mapping.get("trade_date"))) or datetime.utcnow()
    entry_price = parse_float(get(mapping.get("entry_price")))
    exit_price = parse_float(get(mapping.get("exit_price")))
    position_size = parse_float(get(mapping.get("position_size")))
    pnl_amount = parse_float(get(mapping.get("pnl_amount")))
    notes = _stringify(get(mapping.get("notes"))) or None

    # Compute PnL if missing and data is available
    if (
        pnl_amount is None
        and entry_price is not None
        and exit_price is not None
        and position_size is not None
    ):
        pnl = (exit_price - entry_price) * position_size
        pnl_amount = -pnl if direction == "short" else pnl

    payload: Dict[str, Any] = {
        "project_id": project_id,
        "symbol": symbol[:20],
        "direction": direction,
        "trade_date": trade_date,
        "entry_price": entry_price,
        "exit_price": exit_price,
        "position_size": position_size,
        "pnl_amount": pnl_amount,
        "notes": notes,
    }

    # Determine result from pnl_amount when present
    if pnl_amount is not None:
        if pnl_amount > 0:
            payload["result"] = "win"
        elif pnl_amount < 0:
            payload["result"] = "loss"
        else:
            payload["result"] = "breakeven"

    # Drop Nones so SQLAlchemy defaults can apply
    payload = {k: v for k, v in payload.items() if v is not None}

    return payload, None
